import io
import asyncpg
from fastapi import FastAPI, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.worksheet.datavalidation import DataValidation

# -----------------------------
# CONFIG
# -----------------------------
DATABASE_URL = "postgresql://postgres:Bcast%40123@164.52.219.253:5432/sp_dev"

# -----------------------------
# FASTAPI SETUP
# -----------------------------
app = FastAPI(title="Excel Template Generator")


class ExportRequest(BaseModel):
    table: str


# Define your headers (adjust based on your actual structure)
headers = [
    "projectCode",
    "mainActivityCode", "mainActivityName",
    "subActivityCode", "subActivityName",
    "activityCode", "activityName",
    "workGroup",
    "isDeleted",
    "subcontractorActivity",
    "unitSymbol",
    "unitRate",
    "producitivityPerManDay"
]
TASK_COUNT = 3
CATEGORY_COUNT = 3


def get_headers():
    task_headers = []
    category_headers = []

    # Generate task headers
    task_headers = []
    for i in range(TASK_COUNT):
        if i == 0:
            task_headers.extend(['PrimeTaskName', 'PrimeTaskWeightage'])
        else:
            task_headers.extend([f'SubTask{i}Name', f'SubTask{i}Weightage'])

    for i in range(CATEGORY_COUNT):
        category_headers.extend([
            f'WorkmenCategory{i + 1}Name',
            f'WorkmenCategory{i + 1}Hrs',
            f'WorkmenCategory{i + 1}RatePerHour',
        ])

    return [
        *headers,
        'categoryDetails',
        *category_headers,
        'taskDetails',
        *task_headers
    ]


async def get_template_config():
    """Fetch lookup data from database"""
    conn = await asyncpg.connect(DATABASE_URL)
    try:
        # Fetch all lookup data
        projects = await conn.fetch(
            "SELECT project_code FROM projects WHERE is_deleted = false;"
        )
        tasks = await conn.fetch(
            "SELECT task_name FROM tasks WHERE is_deleted = false;"
        )
        categories = await conn.fetch(
            "SELECT category FROM categories WHERE is_deleted = false;"
        )
        units = await conn.fetch(
            "SELECT unit_name FROM units WHERE is_deleted = false;"
        )

        return {
            'projects': [row['project_code'] for row in projects],
            'tasks': [row['task_name'] for row in tasks],
            'categories': [row['category'] for row in categories],
            'units': [row['unit_name'] for row in units]
        }
    finally:
        await conn.close()


def get_column_letter(col_idx):
    """Convert column index to Excel column letter (0-based index)"""
    result = ""
    col_idx += 1  # Convert to 1-based
    while col_idx > 0:
        col_idx -= 1
        result = chr(col_idx % 26 + 65) + result
        col_idx //= 26
    return result


def create_lookup_sheet(wb, sheet_name, column_name, data):
    """Create a hidden lookup sheet with data"""
    if len(data) == 0:
        return '""'

    ws = wb.create_sheet(sheet_name)
    ws.append([column_name])
    for item in data:
        ws.append([item])

    ws.sheet_state = 'veryHidden'

    # Return formula reference
    formula = f"{sheet_name}!$A$2:$A${len(data) + 1}"
    return formula


def create_excel_template(lookup_data, prefill_rows=1000):
    """Create Excel template with dropdowns and formatting"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Add headers
    ws.append(get_headers())

    # Create lookup sheets and get formulas
    project_formula = create_lookup_sheet(
        wb, "ProjectLookup", "projectCode", lookup_data['projects']
    )
    task_formula = create_lookup_sheet(
        wb, "TaskLookup", "taskName", lookup_data['tasks']
    )
    category_formula = create_lookup_sheet(
        wb, "CategoryLookup", "category", lookup_data['categories']
    )
    unit_formula = create_lookup_sheet(
        wb, "UnitLookup", "unitName", lookup_data['units']
    )

    # Add empty rows for data entry
    for _ in range(prefill_rows):
        ws.append([""] * len(get_headers()))

        total_rows = prefill_rows + 1

        # Configure dropdowns
        dropdowns_config = [
            {"header": "projectCode", "formula": project_formula},
            {"header": "workGroup", "values": ["Direct", "Indirect"]},
            {"header": "isDeleted", "values": ["Yes", "No"]},
            {"header": "subcontractorActivity", "values": ["Yes", "No"]},
            {"header": "unitName", "formula": unit_formula},
            {"header": "PrimeTaskName", "formula": task_formula},
            {"header": "SubTask1Name", "formula": task_formula},
            {"header": "SubTask2Name", "formula": task_formula},
            {"header": "SubTask3Name", "formula": task_formula},
            {"header": "WorkmenCategory1Name", "formula": category_formula},
            {"header": "WorkmenCategory2Name", "formula": category_formula},
            {"header": "WorkmenCategory3Name", "formula": category_formula},
        ]

        # Apply data validations
        for dropdown in dropdowns_config:
            header_name = dropdown["header"]
        try:
            col_idx = get_headers().index(header_name)
            col_letter = get_column_letter(col_idx)

            dv = DataValidation(type="list", allow_blank=True)

            if "formula" in dropdown:
                dv.formula1 = dropdown["formula"]
            elif "values" in dropdown:
                dv.formula1 = f'"{",".join(dropdown["values"])}"'

            dv.add(f"{col_letter}2:{col_letter}{total_rows}")
            ws.add_data_validation(dv)
        except ValueError:
            raise
        # Format divider columns
        try:
            task_divider_idx = get_headers().index("taskDetails")
            category_divider_idx = get_headers().index("categoryDetails")

            yellow_fill = PatternFill(
                start_color="FFFFFF00",
                end_color="FFFFFF00",
                fill_type="solid"
            )
            top_align = Alignment(vertical="top", wrap_text=True)

            for divider_idx in [task_divider_idx, category_divider_idx]:
                col_letter = get_column_letter(divider_idx)

                # Merge cells
                ws.merge_cells(f"{col_letter}1:{col_letter}{total_rows}")

                # Apply formatting
                cell = ws[f"{col_letter}1"]
                cell.fill = yellow_fill
                cell.alignment = top_align

                # Set column width
                ws.column_dimensions[col_letter].width = 2
        except ValueError:
            raise

        for row in range(2, total_rows + 1):
            # Format weightage columns as percentage
            for idx, header in enumerate(get_headers()):
                if header.endswith("Weightage"):
                    col_letter = get_column_letter(idx)
                    ws[f"{col_letter}{row}"].number_format = '0.00%'

            # Add formula to unitRate column
            try:
                unit_rate_idx = get_headers().index("unitRate")
                unit_rate_col = get_column_letter(unit_rate_idx)

                productivity_rate_idx = get_headers().index("producitivityPerManDay")
                productivity_col = get_column_letter(productivity_rate_idx)

                # Find the first category name column to calculate range
                category_start_idx = get_headers().index("WorkmenCategory1Name")
                category_start_col = get_column_letter(category_start_idx)

                # Calculate the end column (last category column)
                category_end_idx = category_start_idx + \
                    (CATEGORY_COUNT * 3) - 1
                category_end_col = get_column_letter(category_end_idx)

                # Apply formula to each row
                for row in range(2, total_rows + 1):
                    # Build dynamic formula based on category columns
                    # Extract column letters to variables first
                    category_start_plus1_col = get_column_letter(
                        category_start_idx + 1)
                    category_end_plus1_col = get_column_letter(
                        category_end_idx + 1)
                    productivity_formula = f'''=IFERROR(
                        IF(
                            SUMPRODUCT((MOD(COLUMN({category_start_col}{row}:{category_end_col}{row})-{category_start_idx},3)=1)*IF(ISNUMBER({category_start_col}{row}:{category_end_col}{row}),{category_start_col}{row}:{category_end_col}{row},0))/8=0,
                            "",
                            SUMPRODUCT((MOD(COLUMN({category_start_col}{row}:{category_end_col}{row})-{category_start_idx},3)=1)*IF(ISNUMBER({category_start_col}{row}:{category_end_col}{row}),{category_start_col}{row}:{category_end_col}{row},0))/8
                        ),
                        ""
                    )'''

                    formula = f"""=IFERROR(
                        IF(
                            SUMPRODUCT(
                                (MOD(COLUMN({category_start_col}{row}:{category_end_col}{row}),3)=0)*
                                IF(ISNUMBER({category_start_col}{row}:{category_end_col}{row}),{category_start_col}{row}:{category_end_col}{row},0),
                                (MOD(COLUMN({category_start_plus1_col}{row}:{category_end_plus1_col}{row}),3)=1)*
                                IF(ISNUMBER({category_start_plus1_col}{row}:{category_end_plus1_col}{row}),
                                {category_start_plus1_col}{row}:{category_end_plus1_col}{row},0)
                            )=0,
                            "",
                            SUMPRODUCT(
                                (MOD(COLUMN({category_start_col}{row}:{category_end_col}{row}),3)=0)*
                                IF(ISNUMBER({category_start_col}{row}:{category_end_col}{row}),{category_start_col}{row}:{category_end_col}{row},0),
                                (MOD(COLUMN({category_start_plus1_col}{row}:{category_end_plus1_col}{row}),3)=1)*
                                IF(ISNUMBER({category_start_plus1_col}{row}:{category_end_plus1_col}{row}),
                                {category_start_plus1_col}{row}:{category_end_plus1_col}{row},0)
                            )
                        ),
                        ""
                    )"""

                    ws[f"{unit_rate_col}{row}"] = formula
                    ws[f"{productivity_col}{row}"] = productivity_formula
            except ValueError:
                raise

        # Set header row alignment
        for cell in ws[1]:
            cell.font = Font(bold=True)

        return wb


@app.post("/export")
async def export(req: ExportRequest):
    try:
        # Fetch lookup data from database
        lookup_data = await get_template_config()

        # Create Excel template
        wb = create_excel_template(lookup_data, prefill_rows=1000)

        # Save to buffer
        out_io = io.BytesIO()
        wb.save(out_io)
        out_io.seek(0)

        filename = f"{req.table}_template.xlsx"

        return StreamingResponse(
            out_io,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        print("Unexpected server error:", str(e))
        raise HTTPException(
            status_code=500, detail=f"Internal server error: {str(e)}")

        if __name__ == "__main__":
            import uvicorn
        uvicorn.run(app, host="0.0.0.0", port=8000)
